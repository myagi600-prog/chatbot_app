

import streamlit as st
import google.generativeai as genai
import os
from PIL import Image
import io

# LangChain関連のライブラリをインポート
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_google_genai import GoogleGenerativeAIEmbeddings
from langchain.vectorstores import DocArrayInMemorySearch # FAISSから変更
from langchain.chains.question_answering import load_qa_chain
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain.prompts import PromptTemplate

# ドキュメント読み込み用のライブラリ
from langchain.document_loaders import PyPDFLoader, Docx2txtLoader, TextLoader
import openpyxl

# --- APIキーの設定 ---
try:
    api_key = st.secrets["GEMINI_API_KEY"]
except (FileNotFoundError, KeyError):
    api_key = os.environ.get("GEMINI_API_KEY")
genai.configure(api_key=api_key)

# --- RAG関連の関数 ---
def get_documents_text(uploaded_files):
    text = ""
    for uploaded_file in uploaded_files:
        # 一時ファイルとして保存して、LangChainのLoaderで読み込む
        with open(uploaded_file.name, "wb") as f:
            f.write(uploaded_file.getbuffer())
        
        split_tup = os.path.splitext(uploaded_file.name)
        file_extension = split_tup[1]
        
        try:
            if file_extension == ".pdf":
                loader = PyPDFLoader(uploaded_file.name)
                text += loader.load_and_split()[0].page_content
            elif file_extension == ".docx":
                loader = Docx2txtLoader(uploaded_file.name)
                text += loader.load()[0].page_content
            elif file_extension == ".txt":
                loader = TextLoader(uploaded_file.name)
                text += loader.load()[0].page_content
            elif file_extension in [".xlsx", ".xls"]:
                workbook = openpyxl.load_workbook(uploaded_file.name)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value:
                                text += str(cell.value) + " "
                    text += "\n"
        finally:
            os.remove(uploaded_file.name) # 読み込み後に一時ファイルを削除
    return text

def get_text_chunks(text):
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=200)
    chunks = text_splitter.split_text(text)
    return chunks

def get_vector_store(text_chunks):
    embeddings = GoogleGenerativeAIEmbeddings(model="models/embedding-001")
    # FAISSからDocArrayInMemorySearchに変更
    vector_store = DocArrayInMemorySearch.from_texts(text_chunks, embedding=embeddings)
    return vector_store

def get_conversational_chain():
    prompt_template = """
    以下の文脈情報に基づいて、質問にできるだけ詳しく回答してください。
    文脈:\n {context}?\n
    質問: \n{question}\n
    回答:
    """
    model = ChatGoogleGenerativeAI(model="gemini-1.5-flash", temperature=0.3)
    prompt = PromptTemplate(template=prompt_template, input_variables=["context", "question"])
    chain = load_qa_chain(model, chain_type="stuff", prompt=prompt)
    return chain

# --- Streamlitアプリのメイン部分 ---
st.title("AIチャットボット (RAG機能付き)")
st.caption("サイドバーから知識ファイルをアップロードできます")

# --- サイドバー ---
with st.sidebar:
    st.header("知識ベース設定")
    knowledge_files = st.file_uploader(
        "知識ファイル（PDF/Word/Excel/TXT）をアップロード", 
        accept_multiple_files=True,
        type=["pdf", "docx", "xlsx", "xls", "txt"]
    )
    if st.button("ファイルを処理して知識ベースを構築"):
        if knowledge_files:
            with st.spinner("ファイルを処理中..."):
                raw_text = get_documents_text(knowledge_files)
                text_chunks = get_text_chunks(raw_text)
                # ベクトルストアをセッションステートに保存
                st.session_state.vector_store = get_vector_store(text_chunks)
                st.success("知識ベースの準備ができました！")
        else:
            st.warning("ファイルをアップロードしてください。")

    st.header("接続テスト")
    if st.button("Embedding接続テストを実行"):
        with st.spinner("Embedding APIに接続中..."):
            try:
                embeddings = GoogleGenerativeAIEmbeddings(model="models/embedding-001")
                test_vector = embeddings.embed_query("Hello World")
                st.success("Embedding接続に成功しました！")
                st.write(f"生成されたベクトル（先頭5件）: {test_vector[:5]}")
            except Exception as e:
                st.error(f"Embedding接続中にエラーが発生しました:")
                st.exception(e)

# --- メインチャット画面 ---
if "messages" not in st.session_state:
    st.session_state.messages = []

for message in st.session_state.messages:
    with st.chat_message(message["role"]):
        st.markdown(message["content"])

if prompt := st.chat_input("メッセージを入力してください..."):
    st.session_state.messages.append({"role": "user", "content": prompt})
    with st.chat_message("user"):
        st.markdown(prompt)

    with st.chat_message("assistant"):
        with st.spinner("AIが考えています..."):
            # 知識ベースが準備できている場合
            if "vector_store" in st.session_state:
                try:
                    vector_store = st.session_state.vector_store
                    docs = vector_store.similarity_search(prompt)
                    chain = get_conversational_chain()
                    response = chain({"input_documents": docs, "question": prompt}, return_only_outputs=True)
                    st.markdown(response["output_text"])
                    st.session_state.messages.append({"role": "assistant", "content": response["output_text"]})
                except Exception as e:
                    st.error(f"エラーが発生しました: {e}")
            # 知識ベースがない場合（通常のチャット）
            else:
                try:
                    model = genai.GenerativeModel('gemini-1.5-flash')
                    response = model.generate_content(prompt)
                    st.markdown(response.text)
                    st.session_state.messages.append({"role": "assistant", "content": response.text})
                except Exception as e:
                    st.error(f"エラーが発生しました: {e}")
